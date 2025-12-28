"""
用户批量导入视图
"""
from flask import redirect, url_for, request, flash, send_file
from flask_admin import BaseView, expose
from flask_login import current_user
import os
import tempfile
import io

from flask_app.models import User, Dictionary
from flask_app import db


class UserImportView(BaseView):
    """用户批量导入视图"""
    
    @expose('/', methods=['GET', 'POST'])
    def index(self, cls=None, *args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('auth.login', next=request.url))
        
        if current_user.role != 'admin':
            flash('只有管理员可以访问此页面', 'warning')
            return redirect(url_for('admin.index'))
        
        result = None
        
        if request.method == 'POST':
            file = request.files.get('excel_file')
            
            if file and file.filename.endswith(('.xlsx', '.xls')):
                # 保存临时文件
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    file.save(tmp.name)
                    tmp_path = tmp.name
                
                try:
                    # 导入用户
                    import pandas as pd
                    
                    df = pd.read_excel(tmp_path)
                    
                    success_count = 0
                    skipped_count = 0
                    errors = []
                    users_to_add = []  # 待添加的用户列表
                    
                    # 角色映射
                    role_map = {
                        '学生': 'student',
                        '教师': 'teacher',
                        '老师': 'teacher',
                        '管理员': 'admin',
                        '教学秘书': 'secretary',
                        'student': 'student',
                        'teacher': 'teacher',
                        'admin': 'admin',
                        'secretary': 'secretary'
                    }
                    
                    # 第一阶段：验证所有数据
                    for idx, row in df.iterrows():
                        row_num = idx + 2  # Excel行号（从第2行开始）
                        try:
                            # 获取字段
                            account_id = str(row.get('工号/学号', row.get('学号', row.get('工号', '')))).strip()
                            name = str(row.get('姓名', '')).strip()
                            department = str(row.get('学院', '')).strip()
                            role_str = str(row.get('角色', 'student')).strip()
                            email = str(row.get('邮箱', '')).strip()
                            
                            # 角色转换
                            role = role_map.get(role_str, 'student')
                            
                            # 清理account_id（去除.0后缀）
                            if account_id.endswith('.0'):
                                account_id = account_id[:-2]
                            
                            # 验证必填字段
                            if not account_id or account_id == 'nan':
                                errors.append(f"第{row_num}行: 工号/学号为空")
                                continue
                            
                            if not name or name == 'nan':
                                errors.append(f"第{row_num}行: 姓名为空")
                                continue
                            
                            # 验证角色（必填）
                            if not role_str or role_str == 'nan':
                                errors.append(f"第{row_num}行: 角色为空")
                                continue
                            
                            if role_str not in role_map:
                                errors.append(f"第{row_num}行: 角色'{role_str}'无效，请使用：学生/教师/教学秘书/管理员")
                                continue
                            
                            # 验证学院（必填）
                            if not department or department == 'nan':
                                errors.append(f"第{row_num}行: 学院为空")
                                continue
                            
                            # 验证邮箱（必填）
                            if not email or email == 'nan':
                                errors.append(f"第{row_num}行: 邮箱为空")
                                continue
                            
                            # 检查账号是否已存在
                            existing = User.query.filter_by(account_id=account_id).first()
                            if existing:
                                skipped_count += 1
                                continue
                            
                            # 检查邮箱是否已存在
                            existing_email = User.query.filter_by(email=email).first()
                            if existing_email:
                                errors.append(f"第{row_num}行: 邮箱'{email}'已被使用")
                                continue
                            
                            # 验证通过，添加到待导入列表
                            users_to_add.append({
                                'row_num': row_num,
                                'account_id': account_id,
                                'name': name,
                                'department': department,
                                'role': role,
                                'email': email
                            })
                            
                        except Exception as e:
                            errors.append(f"第{row_num}行: 数据格式错误 - {str(e)}")
                    
                    # 第二阶段：如果有错误，回滚并返回错误信息
                    if errors:
                        db.session.rollback()
                        result = {
                            'success': 0,
                            'skipped': skipped_count,
                            'failed': len(errors),
                            'errors': errors
                        }
                        flash(f"❌ 导入失败，发现 {len(errors)} 个错误，请修正后重新导入", 'danger')
                    else:
                        # 第三阶段：所有验证通过，执行导入
                        try:
                            for user_data in users_to_add:
                                password = user_data['account_id'][-6:] if len(user_data['account_id']) >= 6 else user_data['account_id']
                                
                                user = User(
                                    account_id=user_data['account_id'],
                                    name=user_data['name'],
                                    department=user_data['department'],
                                    role=user_data['role'],
                                    email=user_data['email'],
                                    is_active=True,
                                    created_by=current_user.name
                                )
                                user.set_password(password)
                                db.session.add(user)
                                success_count += 1
                            
                            db.session.commit()
                            
                            result = {
                                'success': success_count,
                                'skipped': skipped_count,
                                'failed': 0,
                                'errors': []
                            }
                            flash(f"✅ 导入成功：新增 {success_count} 个用户，跳过 {skipped_count} 个已存在用户", 'success')
                            
                        except Exception as e:
                            db.session.rollback()
                            result = {
                                'success': 0,
                                'skipped': skipped_count,
                                'failed': len(users_to_add),
                                'errors': [f"导入过程中发生错误：{str(e)}，所有数据已回滚"]
                            }
                            flash(f"❌ 导入失败，数据已回滚：{str(e)}", 'danger')
                    
                except Exception as e:
                    flash(f"❌ 导入出错：{str(e)}", 'danger')
                finally:
                    # 删除临时文件
                    if os.path.exists(tmp_path):
                        os.unlink(tmp_path)
            else:
                flash('请上传 Excel 文件（.xlsx 或 .xls）', 'warning')
        
        return self.render('admin/custom/user_import.html', result=result)
    
    @expose('/download_template')
    def download_template(self):
        """下载用户导入模板（带下拉选择）"""
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = '用户导入'
        
        # 创建隐藏的选项工作表（用于存储下拉选项）
        ws_options = wb.create_sheet('选项数据')
        ws_options.sheet_state = 'hidden'
        
        # 角色选项
        role_list = ['学生', '教师', '管理员', '教学秘书']
        for idx, role in enumerate(role_list, 1):
            ws_options.cell(row=idx, column=1, value=role)
        
        # 学院选项
        college_list = Dictionary.get_values('学院')
        if not college_list:
            college_list = ['计算机学院', '信息工程学院', '电子工程学院']
        
        for idx, college in enumerate(college_list, 1):
            ws_options.cell(row=idx, column=2, value=college)
        
        # 设置标题行
        headers = ['工号/学号', '姓名', '角色', '学院', '邮箱']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # 添加示例数据
        sample_data = [
            ['2023116000001', '张三', '学生', college_list[0] if college_list else '计算机学院', 'zhangsan@example.com'],
            ['10301001', '李老师', '教师', college_list[1] if len(college_list) > 1 else '信息工程学院', 'liteacher@example.com'],
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
        
        # 设置列宽
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 25
        
        # 为角色列添加下拉选择（必填）- 引用隐藏工作表
        role_validation = DataValidation(
            type='list',
            formula1=f'=选项数据!$A$1:$A${len(role_list)}',
            allow_blank=False
        )
        role_validation.error = '请从下拉列表中选择'
        role_validation.errorTitle = '无效输入'
        role_validation.prompt = '请选择角色（必填）'
        role_validation.promptTitle = '角色选择'
        ws.add_data_validation(role_validation)
        role_validation.add('C2:C1000')
        
        # 为学院列添加下拉选择（必填）- 引用隐藏工作表
        college_validation = DataValidation(
            type='list',
            formula1=f'=选项数据!$B$1:$B${len(college_list)}',
            allow_blank=False
        )
        college_validation.error = '请从下拉列表中选择'
        college_validation.errorTitle = '无效输入'
        college_validation.prompt = '请选择学院（必填）'
        college_validation.promptTitle = '学院选择'
        ws.add_data_validation(college_validation)
        college_validation.add('D2:D1000')
        
        # 添加说明工作表
        ws_help = wb.create_sheet('填写说明')
        help_content = [
            ['字段名称', '说明', '是否必填'],
            ['工号/学号', '用户唯一标识，学生填学号，教师填工号', '必填'],
            ['姓名', '用户真实姓名', '必填'],
            ['角色', '下拉选择：学生/教师/教学秘书/管理员', '必填'],
            ['学院', '下拉选择所属学院', '必填'],
            ['邮箱', '用户邮箱，不可重复', '必填'],
            ['', '', ''],
            ['注意事项：', '', ''],
            ['1. 第一行为标题行，数据从第二行开始填写', '', ''],
            ['2. 所有字段均为必填，不可留空', '', ''],
            ['3. 工号/学号和邮箱不能重复，已存在的用户将被跳过', '', ''],
            ['4. 默认密码为工号/学号后6位', '', ''],
        ]
        
        for row_idx, row_data in enumerate(help_content, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_help.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True)
        
        ws_help.column_dimensions['A'].width = 15
        ws_help.column_dimensions['B'].width = 40
        ws_help.column_dimensions['C'].width = 10
        
        # 导出到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='user_import_template.xlsx'
        )
    
    def is_accessible(self):
        return current_user.is_authenticated and current_user.role == 'admin'
